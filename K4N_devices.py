import json
import logging
import requests
from openpyxl import load_workbook
#-----------------------LOGGING/
class ColoredFormatter(logging.Formatter):
    COLORS = {'DEBUG': '\033[95m', 'INFO': '\033[94m', 'WARNING': '\033[93m',
              'ERROR': '\033[91m', 'CRITICAL': '\033[95m'}

    def format(self, record):
        log_fmt = f"%(asctime)s ->>  {self.COLORS.get(record.levelname, '')} %(message)s\033[0m"
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s', handlers=[logging.StreamHandler()])
logging.getLogger().handlers[0].setFormatter(ColoredFormatter())
logging.info('script for K4N - RV SOAR')
#-----------------------LOGGING\
PATH_TO_CERT_K4N = r'C:\Users\d.demin\Desktop\certs\server.crt' # Вставить нужные пути
PATH_TO_KEY_K4N = r'C:\Users\d.demin\Desktop\certs\key.pem' # Вставить нужные пути
ADDR_K4N = '10.22.71.45:8080' # Адрес с портом
#---------------------------


def get_token(addr: str, cert:str, key:str) -> str:
    try:
        logging.info('connected to K4N')
        requests.packages.urllib3.disable_warnings()
        session = requests.Session()
        info = session.post(url= f"https://{addr}/kics/api/auth/v4/token?grant_type=certificate",
                            cert= (cert, key),
                            verify= False)
        logging.info('token получен')
        rez = info.json()

        return rez['access_token']

    except Exception as err:
        logging.error(err)
        logging.warning("no connect, no information")

def get_information_about_product(token: str) -> dict:
    try:
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}"
        }
        requests.packages.urllib3.disable_warnings()
        session = requests.Session()
        info = session.get(url="https://10.22.71.45:8080/kics/api/v3/about",
                                headers= headers,
                                verify=False)
        logging.info("info получена")
        rez = info.json()
        return rez

    except Exception as err:
        logging.error(err)
        logging.warning("no information about product")

def get_unknown_devices(token: str) -> dict:
    try:
        body = {
            "filter": [{"field": "status",
                        "condition": "=",
                        "value": "Unauthorized"}],
            "sort": [],
            "offset": 0, #с какого выгружать
            "limit": 1000 #сколько выгружать
        }
        headers = {
            "Content-Type": "application/json-patch+json",
            "Authorization": f"Bearer {token}"
        }
        requests.packages.urllib3.disable_warnings()
        session = requests.Session()
        response = requests.post(url="https://10.22.71.45:8080/kics/api/v4/devices/query",
                                data=json.dumps(body),
                                headers=headers,
                                verify=False)
        logging.info("devices получены")
        rez = response.json()
        return rez
    except Exception as err:
        logging.error(err)
        logging.warning("no information about devices")

def get_all_events(token: str) -> dict:
    try:
        body = {
            "filter": [{"field": "eventType",
                        "condition": "=",
                        "value": "4000005003"}],
            "sort": [],
            "offset": 0, #с какого выгружать
            "limit": 1000 #сколько выгружать
        }
        headers = {
            "Content-Type": "application/json-patch+json",
            "Authorization": f"Bearer {token}"
        }
        requests.packages.urllib3.disable_warnings()
        session = requests.Session()
        response = requests.post(url="https://10.22.71.45:8080/kics/api/v4/events/query",
                                data=json.dumps(body),
                                headers=headers,
                                verify=False)
        logging.info("events получены")
        rez = response.json()
        return rez
    except Exception as err:
        logging.error(err)
        logging.warning("no information about devices")

def create_xls_devices():
    devices = get_unknown_devices(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))
    wb = load_workbook('Devices import template.xlsx')
    ws = wb["Оборудование"]
    counter = 0
    for i in devices['values']:
        if devices['values'][counter]['name'] == '':
            ws[f'A{counter + 2}'] = 'Имя не найдено'#имя устройства
        else:
            ws[f'A{counter + 2}'] = devices['values'][counter]['name']

        ws[f'B{counter+2}'] = 'Нераспознанный узел' #тип узла

        try:
            ip = [] #для 2х и более ip адресов. R-Vision не добавляет 2 и более ip
            for i in devices['values'][counter]['addressInformation'][0]['ipAddresses']:
                ip.append(i['ip'])
                format = ', '.join(ip)
            ws[f'C{counter+2}'] = format # айпи адрес
        except:
            ws[f'C{counter + 2}'] = ''

        if devices['values'][counter]['addressInformation'][0]['macAddress'] == '':
             ws[f'E{counter + 2}'] = ''# мак адрес
        else:
            ws[f'E{counter + 2}'] = devices['values'][counter]['addressInformation'][0]['macAddress']

        if devices['values'][counter]['os'] == '':
            ws[f'M{counter + 2}'] = 'ОС неизвестна'
        else:
            ws[f'M{counter + 2}'] = devices['values'][counter]['os']

        if devices['values'][counter]['networkName'] == '':
            ws[f'T{counter + 2}'] = 'Сетевое имя неизвестно'
        else:
            ws[f'T{counter + 2}'] = devices['values'][counter]['networkName']
        counter+=1
    wb.save('Devices import template.xlsx')
    wb.close()

def clear_shit():

    wb = load_workbook(filename='Devices import template.xlsx')
    sheet = wb["Оборудование"]
    sheet.delete_rows(2, 1000)
    wb.save('Devices import template.xlsx')

def TESTS():
    #Для девайсов

    for i in get_unknown_devices(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))['values']:
        logging.warning(f'ID: {i['id']} '
              f'IP ADDR: {i['addressInformation'][0]['ipAddresses']}, '
              f'HOSTNAME: {i['name']}, '
              f'MAC ADDR: {i['addressInformation'][0]['macAddress']}')
    '''
    for i in get_all_events(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))['values']:
        logging.warning(f'MAC_ADDR: {i['communications'][0]['sourceMac']} '
              f'IP_ADDR: {i['communications'][0]['sourceIp']}')
    '''
    #logging.error(get_all_events(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N)))
    logging.error(get_unknown_devices(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N)))
    logging.error(len(get_unknown_devices(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))['values']))
    logging.error(len(get_all_events(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))['values']))

def write_in_file():
    data = get_all_events(get_token(ADDR_K4N, PATH_TO_CERT_K4N, PATH_TO_KEY_K4N))
    with open('data.json', 'w') as file:
        json.dump(data, file)

clear_shit()
create_xls_devices()
TESTS()